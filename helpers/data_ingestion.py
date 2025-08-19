import csv
from collections import OrderedDict
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

def read_single_series_bar(path: Path) -> OrderedDict:
    """
    Read a CSV, or an Excel sheet (prefer 'Table1', else first sheet)
    → OrderedDict[category: float]
    """
    suffix = path.suffix.lower()

    if suffix in {'.xls', '.xlsx'}:
        # Try pandas first (fast, uses dtype=str)
        try:
            xls = pd.ExcelFile(path, engine='openpyxl')
            sheets = xls.sheet_names
        except Exception:
            sheets = []

        # If pandas failed to detect sheets, use openpyxl
        if not sheets:
            wb = load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames

        if not sheets:
            raise ValueError(f"No worksheets found in {path.name}")

        # pick the right sheet
        sheet = "Table1" if "Table1" in sheets else sheets[0]
        if sheet != "Table1":
            print(f"[INFO] 'Table1' not found in {path.name}, using '{sheet}' instead.")

        # now read just the first two rows
        wb = wb if 'wb' in locals() else load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
        headers, data_row = rows

        cats = list(headers[1:])
        vals = list(data_row[1:])

    elif suffix == '.csv':
        with path.open('r', encoding='utf-8-sig') as f:
            reader  = csv.reader(f)
            headers = next(reader)
            data    = next(reader)
        cats, vals = headers[1:], data[1:]

    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # normalize to floats
    out = OrderedDict()
    for cat, v in zip(cats, vals):
        s = str(v).strip()
        out[cat.strip()] = float(s.rstrip('%')) if s.endswith('%') else float(s)
    return out